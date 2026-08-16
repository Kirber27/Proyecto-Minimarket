#!/usr/bin/env python3
"""Extrae el catalogo real desde CONTROL DE VENTAS.xlsx hacia mock/*.json"""
import openpyxl, json, re, unicodedata, os

SRC = '/Users/kirberpineda/Downloads/CONTROL DE VENTAS - copia.xlsx'
OUT = '/Users/kirberpineda/Documents/GitHub/Proyecto-Minimarket/mock'
TASA = 800.0  # celda G2 de BODEGA/CHUCHERIA

# Los encabezados de fila del Excel que NO son categorias sino marcas de mes,
# y el bloque sin titulo (fila 162 de BODEGA) que agrupa higiene y limpieza.
CAT_MAP = {
    ('BODEGA', 'ENERO'): 'Viveres',
    ('BODEGA', 'GRANOS'): 'Granos',
    ('BODEGA', 'ENLATADO'): 'Enlatados',
    ('BODEGA', 'CONDIMENTOS'): 'Condimentos',
    ('BODEGA', 'HUEVOS-CHARCUTERIA-POLLO-CHULETA'): 'Charcuteria y proteinas',
    ('BODEGA', 'BEBIDAS'): 'Bebidas',
    ('BODEGA', ''): 'Higiene y limpieza',
    ('BODEGA', 'OTROS ARTICULOS'): 'Otros articulos',
    ('CHUCHERIA', 'AGOSTO'): 'Varios',
    ('CHUCHERIA', 'GALLETAS'): 'Galletas',
    ('CHUCHERIA', 'WAFER'): 'Wafers',
    ('CHUCHERIA', 'SNACK'): 'Snacks',
    ('CHUCHERIA', 'CHOCOLATE'): 'Chocolates',
    ('CHUCHERIA', 'CARAMELOS'): 'Caramelos',
    ('CHUCHERIA', 'CHICLE'): 'Chicles',
    ('CHUCHERIA', 'GOMITAS'): 'Gomitas',
    ('CHUCHERIA', 'TORTAS'): 'Tortas y panques',
}

# Productos que se venden por peso segun el Excel (cantidad decimal / rubro).
PESO = re.compile(r'\b(QUESO|JAMON|MORTADELA|POLLO|CHULETA|CARNE MOLIDA|BISTEC|SALCHICHA)\b')


def slug(s):
    s = unicodedata.normalize('NFKD', s).encode('ascii', 'ignore').decode()
    return re.sub(r'-+', '-', re.sub(r'[^a-z0-9]+', '-', s.lower())).strip('-')


def read_sheet(wb, name):
    ws = wb[name]
    out, cur = [], None
    for r in range(3, ws.max_row + 1):
        a = ws.cell(r, 1).value
        b = ws.cell(r, 2).value
        if a and not b:
            cur = str(a).strip()
            continue
        if not b:
            continue
        nombre = re.sub(r'\s+', ' ', str(b)).strip()
        cantidad = ws.cell(r, 3).value
        precio = ws.cell(r, 5).value
        bs = ws.cell(r, 6).value
        if precio is None:
            continue
        out.append({
            'hoja': name, 'fila': r, 'cat_raw': cur if cur is not None else '',
            'nombre': nombre, 'cantidad': cantidad,
            'precio_usd': round(float(precio), 4), 'bs_excel': bs,
        })
    return out


def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    raw = read_sheet(wb, 'BODEGA') + read_sheet(wb, 'CHUCHERIA')

    categorias, productos, avisos = {}, [], []
    vistos = {}
    for i, p in enumerate(raw, 1):
        cat = CAT_MAP.get((p['hoja'], p['cat_raw']))
        if cat is None:
            avisos.append(f"categoria sin mapear: {p['hoja']}!{p['cat_raw']!r} fila {p['fila']}")
            cat = 'Sin categoria'
        unidad = 'KG' if PESO.search(p['nombre']) else 'UND'
        stock = float(p['cantidad']) if isinstance(p['cantidad'], (int, float)) else 0.0

        # Detecta el error de formula de la hoja: CAMBIO BS. debe ser precio * TASA.
        esperado = round(p['precio_usd'] * TASA, 2)
        if isinstance(p['bs_excel'], (int, float)) and abs(p['bs_excel'] - esperado) > 1:
            avisos.append(
                f"{p['hoja']}!fila {p['fila']} {p['nombre']}: CAMBIO BS. = {p['bs_excel']} "
                f"pero precio {p['precio_usd']} x {int(TASA)} = {esperado}")

        base = slug(p['nombre'])
        vistos[base] = vistos.get(base, 0) + 1
        sku = base if vistos[base] == 1 else f'{base}-{vistos[base]}'

        categorias.setdefault(cat, {
            'id': slug(cat), 'nombre': cat, 'unidad_negocio': 'bodega',
            'origen_hoja': p['hoja'], 'orden': len(categorias) + 1,
        })
        productos.append({
            'id': f'prd-{i:04d}', 'sku': sku, 'nombre': p['nombre'].title(),
            'nombre_original': p['nombre'], 'categoria_id': slug(cat),
            'unidad_negocio': 'bodega', 'unidad_medida': unidad,
            'precio_venta_usd': p['precio_usd'],
            'costo_usd': None,  # el Excel no registra precio de compra
            'stock_actual': stock, 'stock_minimo': 5 if unidad == 'UND' else 1,
            'activo': True, 'origen': f"{p['hoja']}!B{p['fila']}",
        })

    # --- Clientes y deudas (hoja DEUDAS 2026) ---
    ws = wb['DEUDAS 2026']
    clientes, deudas = [], []
    for r in range(4, ws.max_row + 1):
        nombre = ws.cell(r, 2).value
        if not nombre:
            continue
        nombre = re.sub(r'\s+', ' ', str(nombre)).strip()
        cid = f'cli-{len(clientes) + 1:03d}'
        clientes.append({'id': cid, 'nombre': nombre.title(), 'telefono': None,
                         'activo': True, 'origen': f'DEUDAS 2026!B{r}'})
        for col, unidad, moneda in ((3, 'bodega', None), (5, 'bodega', 'USD'),
                                    (6, 'cerveza', 'VES'), (7, 'thais', 'USD')):
            v = ws.cell(r, col).value
            if v in (None, ''):
                continue
            if isinstance(v, (int, float)):
                deudas.append({'cliente_id': cid, 'unidad_negocio': unidad,
                               'monto': float(v), 'moneda': moneda or 'VES',
                               'nota_original': None, 'requiere_revision': False,
                               'origen': f'DEUDAS 2026!{chr(64 + col)}{r}'})
            else:
                # Texto libre tipo "4,5+1,80+1refresco": no es convertible a monto.
                deudas.append({'cliente_id': cid, 'unidad_negocio': unidad,
                               'monto': None, 'moneda': 'USD',
                               'nota_original': str(v).strip(), 'requiere_revision': True,
                               'origen': f'DEUDAS 2026!{chr(64 + col)}{r}'})

    denominaciones = {
        'VES': [5, 10, 20, 50, 100, 200, 500],
        'USD': [1, 5, 10, 20, 50, 100],
    }
    unidades_negocio = [
        {'id': 'bodega', 'nombre': 'Bodega', 'activo': True},
        {'id': 'cerveza', 'nombre': 'Cerveza', 'activo': True},
        {'id': 'thais', 'nombre': 'Thais', 'activo': True},
    ]
    metodos_pago = [
        {'id': 'efectivo-ves', 'nombre': 'Efectivo Bs.', 'moneda': 'VES', 'afecta_arqueo': True},
        {'id': 'efectivo-usd', 'nombre': 'Efectivo $', 'moneda': 'USD', 'afecta_arqueo': True},
        {'id': 'punto', 'nombre': 'Punto de venta', 'moneda': 'VES', 'afecta_arqueo': False},
        {'id': 'pago-movil', 'nombre': 'Pago movil', 'moneda': 'VES', 'afecta_arqueo': False},
        {'id': 'biopago', 'nombre': 'Biopago', 'moneda': 'VES', 'afecta_arqueo': False},
        {'id': 'credito', 'nombre': 'Fiado', 'moneda': 'USD', 'afecta_arqueo': False},
    ]

    os.makedirs(OUT, exist_ok=True)
    dump = {
        'unidades-negocio.json': unidades_negocio,
        'metodos-pago.json': metodos_pago,
        'categorias.json': sorted(categorias.values(), key=lambda c: c['orden']),
        'productos.json': productos,
        'clientes.json': clientes,
        'deudas.json': deudas,
        'denominaciones.json': denominaciones,
        'tasa-cambio.json': [{'fecha': '2026-08-16', 'moneda_base': 'USD',
                              'moneda_destino': 'VES', 'tasa': TASA,
                              'origen': 'BODEGA!G2'}],
    }
    for fname, data in dump.items():
        with open(os.path.join(OUT, fname), 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        print(f'{fname}: {len(data)} registros')

    print(f'\n{len(avisos)} avisos de calidad de datos:')
    for a in avisos:
        print(' -', a)


if __name__ == '__main__':
    main()
